import openpyxl
def readExcel(inputfilename):
    data = []
    wb = openpyxl.load_workbook(inputfilename)
    for ws in wb:
        for wr in ws.rows:
            row = []
            for cell in wr:
                row.append(cell.value)
            data.append(row)
    return data

def writeExcel(outputHeader, outputRows, outputFileName):
    print outputHeader
    print outputRows
    wb = openpyxl.Workbook()
    ws = wb.active
    for j, headerValue in enumerate(outputHeader):
        ws.cell(row = 1, column = j).value = headerValue
    for i, row in enumerate(outputRows):
        for j, value in enumerate(row):
            ws.cell(row = i, column = j).value = value
    wb.save(outputFileName)

def printData(data):
    for row in data:
        for value in row:
            print value,
        print

data = readExcel("input-v3.xlsx")
outputHeader = data[0]
outputRows = data[1:]
printData(outputRows)
writeExcel(outputHeader, outputRows, "input-v3-regions.xlsx")

